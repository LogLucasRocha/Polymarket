"""Persistência dos resultados do backtest.

A cada rodada (a cada 3 dias + watchdog), grava:
  - backtest_results/resultados_backtest.xlsx — planilha Excel com o ÚLTIMO
    resultado de cada estratégia, para o Lucas analisar e tirar insights;
  - docs/generated_numbers.tex — os mesmos números como macros LaTeX, para o
    documento didático (docs/tmax.tex) nunca ficar desatualizado.
O workflow de backtest commita ambos. Nada disso é enviado ao Telegram.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config

RESULTS_DIR = config.ROOT / "backtest_results"
XLSX = RESULTS_DIR / "resultados_backtest.xlsx"
NUMBERS_TEX = config.ROOT / "docs" / "generated_numbers.tex"

_HOUR_WORD = {12: "twelve", 14: "fourteen", 16: "sixteen"}

_HEAD_FILL = PatternFill("solid", fgColor="1A5FB4")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_ACTIVE_FILL = PatternFill("solid", fgColor="D7F0DD")
_PCT = "0.0%"
_MULT = '0.00"x"'


def persist(edge: dict, harvests: dict, combined: dict, conf: dict,
            cal: dict, days: int) -> None:
    """Grava o Excel de resultados e regenera os números do documento."""
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")
    span = max(round(days / max(len(config.STATIONS), 1)), 1)
    _write_xlsx(edge, harvests, combined, conf, cal, days, span, stamp)
    _write_numbers(edge, harvests, combined, conf, span, stamp)


# ------------------------------------------------------------------ Excel

def _header(ws, cols, row=1):
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=row, column=j, value=name)
        c.fill, c.font = _HEAD_FILL, _HEAD_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def _autofit(ws, widths):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _monthly(comp, span):
    return comp ** (30.0 / span) - 1.0 if comp > 0 else 0.0


def _write_xlsx(edge, harvests, combined, conf, cal, days, span, stamp):
    wb = Workbook()

    # ---- Resumo ----
    ws = wb.active
    ws.title = "Resumo"
    _header(ws, ["Estratégia", "Entradas", "Acerto", "P&L flat (×inicial)",
                 "Composto (×)", "Retorno/mês aprox.", "Drawdown máx",
                 "Stops", "Preço médio"])
    active_h = config.HARVEST_MIN_HOUR
    linhas = [("Edge (compra NÃO)", edge, False)]
    for h in sorted(harvests):
        linhas.append((f"Colheita {h}h", harvests[h], h == active_h))
    linhas.append((f"Combinado (Edge + Colheita {active_h}h)", combined, False))

    r = 2
    for nome, st, ativa in linhas:
        if not st.get("n"):
            continue
        vals = [nome, st["n"], st["hit"], st["flat"], st["compounded"],
                _monthly(st["compounded"], span), st["maxdd"],
                st.get("n_stopped", 0), st.get("avg_price", 0)]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v)
            if j in (3, 6, 7):
                c.number_format = _PCT
            elif j == 5:
                c.number_format = _MULT
            elif j in (4, 9):
                c.number_format = "0.00"
            if ativa:
                c.fill = _ACTIVE_FILL
        r += 1
    ws.cell(row=r + 1, column=1,
            value=f"Atualizado em {stamp} · arquivo de {days} dias-cidade "
                  f"(~{span} dias corridos) · {len(config.STATIONS)} cidades")
    ws.cell(row=r + 2, column=1,
            value="Flat = 10% do capital inicial por aposta (compara "
                  "estratégias). Composto = 10% do capital corrente (quanto "
                  "eu teria). In-sample — expectativa executada é menor.")
    _autofit(ws, [30, 10, 9, 17, 13, 17, 13, 8, 11])

    # ---- Por cidade ----
    ws = wb.create_sheet("Por cidade")
    _header(ws, ["Estratégia", "Cidade", "Entradas", "Acerto", "P&L flat"])
    r = 2
    for nome, st in (("Edge", edge), ("Combinado", combined)):
        for icao, (n, w, pnl) in sorted(st.get("by_city", {}).items()):
            ws.cell(row=r, column=1, value=nome)
            ws.cell(row=r, column=2,
                    value=f"{config.STATIONS[icao].city} ({icao})"
                    if icao in config.STATIONS else icao)
            ws.cell(row=r, column=3, value=n)
            c = ws.cell(row=r, column=4, value=(w / n) if n else 0)
            c.number_format = _PCT
            c = ws.cell(row=r, column=5, value=pnl)
            c.number_format = "0.00"
            r += 1
    _autofit(ws, [12, 24, 10, 9, 10])

    # ---- Confiança ----
    ws = wb.create_sheet("Confiança ≥90%")
    _header(ws, ["Cidade", "Faixas-dia", "Acerto real", "Declarado"])
    r = 2
    for icao, g in sorted(conf.get("por_faixa_dia", {}).items()):
        ws.cell(row=r, column=1,
                value=f"{config.STATIONS[icao].city} ({icao})"
                if icao in config.STATIONS else icao)
        ws.cell(row=r, column=2, value=g["n"])
        c = ws.cell(row=r, column=3, value=g["acerto"])
        c.number_format = _PCT
        c = ws.cell(row=r, column=4, value=g["conf_media"])
        c.number_format = _PCT
        r += 1
    _autofit(ws, [24, 12, 12, 11])

    # ---- Calibração ----
    ws = wb.create_sheet("Calibração")
    _header(ws, ["Período", "Brier cru", "Brier calibrado", "Blend: peso "
                 "modelo (a)", "Blend: peso preço (b)", "Brier posterior"])
    r = 2
    for per in ("0-5h", "6-11h", "12-23h"):
        iso = cal.get(per, {})
        bl = cal.get(f"blend {per}", {})
        ws.cell(row=r, column=1, value=per)
        for j, key, src in ((2, "brier_raw", iso), (3, "brier_cal", iso),
                            (4, "a", bl), (5, "b", bl),
                            (6, "brier_post", bl)):
            if key in src:
                ws.cell(row=r, column=j, value=round(src[key], 4))
        r += 1
    _autofit(ws, [10, 11, 15, 18, 18, 15])

    # ---- Meta ----
    ws = wb.create_sheet("Meta")
    for i, (k, v) in enumerate((
            ("Atualizado (UTC)", stamp),
            ("Dias-cidade no arquivo", days),
            ("Dias corridos (aprox.)", span),
            ("Cidades em operação", len(config.STATIONS)),
            ("Colheita ativa (hora local mín.)", config.HARVEST_MIN_HOUR),
            ("Faixas com resolução divergente", edge.get("res_mismatch", 0)),
    ), 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    _autofit(ws, [34, 26])

    wb.save(XLSX)


# ------------------------------------------------------------------ LaTeX

def _write_numbers(edge, harvests, combined, conf, span, stamp):
    lines = ["% Gerado por tmax/results.py a cada backtest — NÃO editar à mão.",
             f"% Atualizado em {stamp}", ""]

    def cmd(name, value):
        lines.append(f"\\newcommand{{\\{name}}}{{{value}}}")

    def pct(x):
        return f"{x * 100:.0f}\\%"

    def strat(prefix, st):
        cmd(f"{prefix}N", st["n"])
        cmd(f"{prefix}Hit", pct(st["hit"]))
        cmd(f"{prefix}Comp", f"{st['compounded']:.2f}")
        cmd(f"{prefix}Flat", f"{st['flat']:+.2f}")
        cmd(f"{prefix}DD", pct(st["maxdd"]))
        cmd(f"{prefix}Stops", st["n_stopped"])
        cmd(f"{prefix}Price", f"{st.get('avg_price', 0):.2f}")

    strat("edge", edge)
    strat("comb", combined)
    cmd("combMonthly", f"{_monthly(combined['compounded'], span) * 100:+.0f}\\%")

    for h, st in harvests.items():
        strat(f"harv{_HOUR_WORD.get(h, 'x')}", st)
    active = harvests.get(config.HARVEST_MIN_HOUR)
    if active:
        strat("harv", active)
    cmd("harvActiveHour", config.HARVEST_MIN_HOUR)

    fd = conf.get("por_faixa_dia", {})
    if fd:
        accs = [v["acerto"] for v in fd.values()]
        cmd("confMin", pct(min(accs)))
        cmd("confMax", pct(max(accs)))

    cmd("btDaysCity", edge.get("days", 0))
    cmd("btSpan", span)
    cmd("btStations", len(config.STATIONS))
    cmd("btUpdated", stamp.replace("T", " ").replace("+00:00", " UTC"))

    NUMBERS_TEX.parent.mkdir(parents=True, exist_ok=True)
    NUMBERS_TEX.write_text("\n".join(lines) + "\n", encoding="utf-8")
